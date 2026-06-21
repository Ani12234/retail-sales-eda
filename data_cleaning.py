import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def clean_data():
    raw_path = "retail_sales_raw.csv"
    if not os.path.exists(raw_path):
        print(f"Error: Raw dataset not found at {raw_path}! Please run data_generation.py first.")
        return
        
    print(f"Loading raw dataset from {raw_path}...")
    df = pd.read_csv(raw_path)
    before_count = len(df)
    print(f"Original record count: {before_count} rows")
    
    # 1. Structural Cleansing
    print("\n--- Step 1: Structural Cleansing ---")
    # Identify duplicates
    duplicates_mask = df.duplicated()
    dup_count = duplicates_mask.sum()
    df = df.drop_duplicates()
    print(f"Removed duplicates: {dup_count} rows")
    
    # Identify negative prices/revenues
    negatives_mask = (df['revenue'] < 0) | (df['unit_price'] < 0)
    neg_count = negatives_mask.sum()
    df = df[~negatives_mask]
    print(f"Removed invalid negative price/revenue records: {neg_count} rows")
    
    structural_removed = dup_count + neg_count
    
    # 2. Outlier Detection using Z-Score method
    print("\n--- Step 2: Outlier Detection (Z-Score) ---")
    mean_rev = df['revenue'].mean()
    std_rev = df['revenue'].std()
    z_rev = (df['revenue'] - mean_rev) / std_rev
    
    mean_units = df['units_sold'].mean()
    std_units = df['units_sold'].std()
    z_units = (df['units_sold'] - mean_units) / std_units
    
    rev_outliers = np.abs(z_rev) > 3.0
    units_outliers = np.abs(z_units) > 3.0
    outliers_mask = rev_outliers | units_outliers
    outliers_count = outliers_mask.sum()
    
    df_clean = df[~outliers_mask].copy()
    print(f"Removed statistical outliers: {outliers_count} rows")
    
    # 3. Missing Values Imputation
    print("\n--- Step 3: Missing Value Imputation ---")
    # Count missing values before imputation
    missing_counts = df_clean.isna().sum()
    print("Missing values detected:")
    for col, count in missing_counts.items():
        if count > 0:
            print(f"  {col}: {count} missing values")
            
    # Calculate imputation metrics
    median_revenue = df_clean['revenue'].median()
    median_units = df_clean['units_sold'].median()
    mode_region = df_clean['region'].mode()[0]
    mode_category = df_clean['product_category'].mode()[0]
    
    # Impute
    df_clean['revenue'] = df_clean['revenue'].fillna(median_revenue)
    df_clean['units_sold'] = df_clean['units_sold'].fillna(median_units).astype(int)
    df_clean['region'] = df_clean['region'].fillna(mode_region)
    df_clean['product_category'] = df_clean['product_category'].fillna(mode_category)
    
    # Ensure datatypes
    df_clean['order_date'] = pd.to_datetime(df_clean['order_date']).dt.date
    df_clean['unit_price'] = np.round(df_clean['revenue'] / df_clean['units_sold'], 2)
    
    print("\nImputation Summary:")
    print(f"  Imputed 'revenue' with median: ${median_revenue:,.2f}")
    print(f"  Imputed 'units_sold' with median: {median_units}")
    print(f"  Imputed 'region' with mode: '{mode_region}'")
    print(f"  Imputed 'product_category' with mode: '{mode_category}'")
    
    # 4. Accuracy & Summary Metrics
    print("\n--- Step 4: Summary Metrics & Accuracy ---")
    after_count = len(df_clean)
    accuracy_score = (after_count / before_count) * 100
    total_imputed = missing_counts.sum()
    
    print(f"Raw Records: {before_count}")
    print(f"Duplicates/Negatives Removed: {structural_removed}")
    print(f"Outliers Removed: {outliers_count}")
    print(f"Missing Values Imputed: {total_imputed}")
    print(f"Clean, Valid Records Remaining: {after_count}")
    print(f"Data Accuracy Score: {accuracy_score:.2f}% (Target: ~97.00%)")
    
    # 5. Export clean datasets
    print("\n--- Step 5: Exporting Datasets ---")
    out_dir = "."
    
    # Export to CSV
    csv_out = os.path.join(out_dir, "retail_sales_clean.csv")
    df_clean.to_csv(csv_out, index=False)
    print(f"Successfully exported clean CSV: {csv_out}")
    
    # Export to styled multi-sheet Excel with Power BI specs
    excel_out = os.path.join(out_dir, "power_bi_dashboard_spec.xlsx")
    export_to_styled_spec_excel(df_clean, excel_out)
    print(f"Successfully exported clean Excel + Power BI Spec: {excel_out}")
    
    return accuracy_score, outliers_count, total_imputed

def export_to_styled_spec_excel(df, file_path):
    wb = Workbook()
    
    # -------------------------------------------------------------
    # SHEET 1: Cleaned Sales Data
    # -------------------------------------------------------------
    ws_data = wb.active
    ws_data.title = "Cleaned Sales Data"
    ws_data.views.sheetView[0].showGridLines = True
    
    # Add headers and rows
    headers = list(df.columns)
    ws_data.append(headers)
    for r in dataframe_to_rows(df, index=False, header=False):
        ws_data.append(r)
        
    # Styling variables
    font_family = "Segoe UI"
    navy_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=11, bold=True)
    normal_font = Font(name=font_family, size=11)
    
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    
    thin_border_side = Side(style='thin', color='D9D9D9')
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Style Sheet 1
    for col_idx in range(1, len(headers) + 1):
        cell = ws_data.cell(row=1, column=col_idx)
        cell.fill = navy_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border
        
    for r_idx in range(2, ws_data.max_row + 1):
        for c_idx in range(1, len(headers) + 1):
            cell = ws_data.cell(row=r_idx, column=c_idx)
            cell.font = normal_font
            cell.border = thin_border
            col_name = headers[c_idx - 1]
            if col_name in ['order_id', 'quarter', 'region', 'product_category', 'customer_segment']:
                cell.alignment = align_left
            elif col_name == 'order_date':
                cell.alignment = align_center
                cell.number_format = 'yyyy-mm-dd'
            elif col_name == 'units_sold':
                cell.alignment = align_right
                cell.number_format = '#,##0'
            elif col_name in ['unit_price', 'revenue']:
                cell.alignment = align_right
                cell.number_format = '$#,##0.00'
                
    # Auto-fit columns for Sheet 1
    for col in ws_data.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws_data.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # -------------------------------------------------------------
    # SHEET 2: Theme Setup & DAX
    # -------------------------------------------------------------
    ws_theme = wb.create_sheet(title="Dashboard Setup & Theme")
    ws_theme.views.sheetView[0].showGridLines = True
    
    theme_headers = ["Parameter", "Value / Color Hex / DAX", "Description"]
    ws_theme.append(theme_headers)
    for col_idx in range(1, 4):
        cell = ws_theme.cell(row=1, column=col_idx)
        cell.fill = navy_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border
        
    theme_rows = [
        ["Theme Setting", "Dark Mode", "Dashboard background should be dark (e.g. #0F172A)"],
        ["Primary Color (Slate Blue)", "#3B82F6", "Used for main visuals, charts, and normal data points"],
        ["Secondary Color (Dark Slate)", "#1E2937", "Used for card backgrounds and container structures"],
        ["Accent Color (Gold)", "#F59E0B", "Used for highlighting the bottom 3 underperforming regions and shortfalls"],
        ["Text Color", "#F3F4F6", "Light color for readability on dark backgrounds"],
        ["Gridlines / Muted Elements", "#374151", "Used for visual borders and line separators"],
        ["Font Family", "Segoe UI / Outfit", "Standard font for clean business presentation"],
        ["Total Revenue DAX Measure", "Total Revenue = SUM('Cleaned Sales Data'[revenue])", "Sum of revenue field"],
        ["Average Regional Revenue Target", "Avg Region Revenue Target = CALCULATE([Total Revenue], ALL('Cleaned Sales Data'[region])) / DISTINCTCOUNT('Cleaned Sales Data'[region])", "Average target per region (10% share)"],
        ["Regional Shortfall DAX", "Regional Shortfall = VAR AvgTarget = [Avg Region Revenue Target] VAR ActualRev = [Total Revenue] RETURN IF(ActualRev < AvgTarget, AvgTarget - ActualRev, 0)", "Shortfall of bottom regions vs average"]
    ]
    
    for row in theme_rows:
        ws_theme.append(row)
        
    for r_idx in range(2, ws_theme.max_row + 1):
        for c_idx in range(1, 4):
            cell = ws_theme.cell(row=r_idx, column=c_idx)
            cell.font = normal_font
            cell.border = thin_border
            if c_idx == 1:
                cell.font = bold_font
                cell.alignment = align_left
            elif c_idx == 2:
                cell.alignment = align_left
                # Highlight hex colors visually
                val = str(cell.value)
                if val.startswith("#"):
                    if "F59E0B" in val:
                        cell.fill = PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid")
                        cell.font = Font(name=font_family, size=11, bold=True, color="B45309")
                    elif "3B82F6" in val:
                        cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
                        cell.font = Font(name=font_family, size=11, bold=True, color="1D4ED8")
                    elif "1E2937" in val:
                        cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
                        cell.font = Font(name=font_family, size=11, bold=True, color="1F2937")
            else:
                cell.alignment = align_left
                
    # Auto-fit Sheet 2
    ws_theme.column_dimensions['A'].width = 35
    ws_theme.column_dimensions['B'].width = 50
    ws_theme.column_dimensions['C'].width = 60
    
    # -------------------------------------------------------------
    # SHEET 3: Visual Specifications
    # -------------------------------------------------------------
    ws_spec = wb.create_sheet(title="Visual Specifications")
    ws_spec.views.sheetView[0].showGridLines = True
    
    spec_headers = ["Visual Name", "Power BI Visual Type", "Field Mapping (Axis / Legend / Values)", "Formatting Details", "Target Metric"]
    ws_spec.append(spec_headers)
    for col_idx in range(1, 6):
        cell = ws_spec.cell(row=1, column=col_idx)
        cell.fill = navy_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border
        
    spec_rows = [
        ["KPI: Total Revenue", "Card", "Fields: [Total Revenue]", "Display unit: Auto, Font: Bold, Color: #F3F4F6", "$4,800,000.00 (before filters)"],
        ["Revenue by Product Category", "Clustered Bar Chart", "Y-Axis: product_category\nX-Axis: [Total Revenue]", "Fill Color: Bright Slate Blue (#3B82F6), Sorted Descending", "Electronics is top seller, Books is lowest"],
        ["Regional Sales Performance", "Clustered Column Chart", "X-Axis: region\nY-Axis: [Total Revenue]", "Conditional Formatting: Fill Color = Accent Gold (#F59E0B) if region is Northwest, Mountain, or New England; else Bright Slate Blue (#3B82F6)", "Bottom 3 regions highlighted, summing to exactly 18.00% of total revenue"],
        ["Quarterly Trend vs. Baseline", "Line Chart", "X-Axis: quarter\nY-Axis: [Total Revenue]", "Line Color: Bright Slate Blue (#3B82F6). Add Constant Line on Y-Axis representing the baseline average of Q1-Q3", "Shows Q4 revenue spike of exactly 23.00% over the baseline average"],
        ["Bottom Regions Shortfall", "Clustered Column Chart (grouped)", "X-Axis: region\nY-Axis: [Total Revenue], [Avg Region Revenue Target]\nVisual Filter: region is Northwest, Mountain, New England", "Highlight [Total Revenue] in Gold, [Avg Region Revenue Target] in Dark Grey/Semi-transparent", "Visually displays the shortfall (totaling 12.00% of total revenue, representing an 18.00% shortfall from target)"]
    ]
    
    for row in spec_rows:
        ws_spec.append(row)
        
    for r_idx in range(2, ws_spec.max_row + 1):
        for c_idx in range(1, 6):
            cell = ws_spec.cell(row=r_idx, column=c_idx)
            cell.font = normal_font
            cell.border = thin_border
            if c_idx == 1:
                cell.font = bold_font
            cell.alignment = align_left
            
    # Auto-fit Sheet 3
    ws_spec.column_dimensions['A'].width = 30
    ws_spec.column_dimensions['B'].width = 25
    ws_spec.column_dimensions['C'].width = 45
    ws_spec.column_dimensions['D'].width = 55
    ws_spec.column_dimensions['E'].width = 40
    
    wb.save(file_path)

if __name__ == "__main__":
    clean_data()
